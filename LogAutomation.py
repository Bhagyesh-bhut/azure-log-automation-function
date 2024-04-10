from azure.identity import DefaultAzureCredential
from azure.monitor.query import LogsQueryClient, MetricsQueryClient
from datetime import datetime, timezone , timedelta
import pandas as pd
from openpyxl import load_workbook,Workbook
from datetime import datetime,timedelta,date
import queries_list
from azure.storage.blob import BlobServiceClient
from openpyxl.styles import Font

today = date.today()
yesterday = today - timedelta(days = 1)

#set the query based on sheet name and query for that particular sheet
# quries={
#   'Alert Hourly Count':queries_list.alert_count,
#   'Financial Account Hourly Count':queries_list.account_count,
#   'Disbursements Hourly Count':queries_list.disbursement_count,
#   'Collateral Hourly Count':queries_list.collateral_count,
#   'Insurance Hourly Count':queries_list.insurance_count,
#   'Customer Hourly Count':queries_list.customer_count,
#   'Customer-error-list':queries_list.customer_error_count,
#   'Alert-error-list':queries_list.alert_error_count,
#   'Financial-Account-error-list':queries_list.account_error_count,
#   'Disbursements-error-list':queries_list.disbursment_error_count,
#   'Collateral-error-list':queries_list.collateral_error_count,
#   'Insurance-error-list':queries_list.insurance_error_count
# }

# log_count_queries={
#   'Alert Hourly Count':queries_list.alert_count,
#   'Financial Account Hourly Count':queries_list.account_count,
#   'Disbursements Hourly Count':queries_list.disbursement_count,
#   'Collateral Hourly Count':queries_list.collateral_count,
#   'Insurance Hourly Count':queries_list.insurance_count,
#   'Customer Hourly Count':queries_list.customer_count
# }

# error_count_queries={
#   'Customer-error-list':queries_list.customer_error_count,
#   'Alert-error-list':queries_list.alert_error_count,
#   'Financial-Account-error-list':queries_list.account_error_count,
#   'Disbursements-error-list':queries_list.disbursment_error_count,
#   'Collateral-error-list':queries_list.collateral_error_count,
#   'Insurance-error-list':queries_list.insurance_error_count
# }

log_count_queries={
  'Alert Hourly Count':'exceptions | take 5 | project appName,severityLevel',
  'Financial Account Hourly Count': 'exceptions | take 1 | project appName,severityLevel'
}

error_count_queries={
  'Alert-error-list':'exceptions | take 5 | project appName',
  'Financial-Account-error-list': 'exceptions | take 1 | project appName'
}



def create_connection():
  #create azure connection 
  credential = DefaultAzureCredential()
  return LogsQueryClient(credential)
 

def execute_queries(log_count_queries,error_count_queries,file_name):

  logs_client=create_connection()

  #run the log queries one by one
  for key, value in log_count_queries.items():
    # print(value)
    sheetName=key
    
    try:
        
      #execte the query in azure app insights logs
      #Please remove timespan parameter from below function if you are using custom timestamp in queries
      response=logs_client.query_resource("subscriptions/c6ac49c8-0c7a-4b98-9d4d-6970b017e7f9/resourceGroups/mule/providers/microsoft.insights/components/azure-app-insight",query=value,timespan=timedelta(days=45))

      # print(response)

      #get the query result
      data=response.tables

      for table in data:
        #convert query result in pandas dataframe
        df = pd.DataFrame(data=table.rows, columns=table.columns)

      #append result in excel sheet
      export_query_result(file_name,key,df,"log_count")



    except Exception as e:
      print("An exception occurred")
      print(e)

    #run the error queries one by one
  for key, value in error_count_queries.items():
    sheetName=key
    
    try:
        
      #execte the query in azure app insights logs
      #Please remove timespan parameter from below function if you are using custom timestamp in queries
      response=logs_client.query_resource("subscriptions/c6ac49c8-0c7a-4b98-9d4d-6970b017e7f9/resourceGroups/mule/providers/microsoft.insights/components/azure-app-insight",query=value,timespan=timedelta(days=45))
      


      #get the query result
      data=response.tables

      for table in data:
        #convert query result in pandas dataframe
        df = pd.DataFrame(data=table.rows, columns=table.columns)

      #append result in excel sheet
      export_query_result(file_name,key,df,"error_count")

    except Exception as e:
      print("An exception occurred")
      print(e)




def export_query_result(file_name,sheetName,data,query_type):
    
   
    container_name = 'logdata'
    blob_name = file_name  # name of the Excel file in Azure Storage

    # Connect to Azure Storage account
    
    blob_service_client = BlobServiceClient.from_connection_string(connection_string)
    container_client = blob_service_client.get_container_client(container_name)

    # Check if the file exists in Azure Blob Storage
    blob_client = container_client.get_blob_client(blob_name)
    blob_exists = blob_client.exists()

    # Load the workbook if the file exists, otherwise create a new workbook
    if blob_exists:
      # Download the Excel file from Azure Storage
      blob_client = container_client.get_blob_client(blob_name)
      with open(blob_name, "wb") as my_blob:
        download_stream = blob_client.download_blob()
        my_blob.write(download_stream.readall())

      workbook = load_workbook(blob_name)
    else:
      workbook = Workbook()

     
      
    

      #check the sheet name in file if sheet is alerady exits then append the result otherwise create new sheet based on the queries(dictionary) key
    if sheetName in workbook.sheetnames:
      worksheet = workbook[sheetName]
    else:
      worksheet = workbook.create_sheet(title=sheetName)

    

    # Calculate total count
    if query_type=="log_count":
        total_amount = data['severityLevel'].sum()
      #total_Id_Count=data['IdCount'].sum()
  
    elif query_type=="error_count":
        #total_error_count=data['responseCode'].count()
        total_error_count=data['appName'].count()


    #get the column List from the dataframe
    column_list=data.columns.tolist()
    #get the value List from the dataframe
    value_list = data.values.tolist()

    #append data after one row
    worksheet.append([])
    #set the previous day date in sheet because queries run for the collect previous day logs
    worksheet.append([yesterday])
    cell = worksheet.cell(row=worksheet.max_row, column=1)
    cell.font = Font(bold=True)

    #append column in sheet
    worksheet.append(column_list)

    #append values in sheet
    #worksheet.append(value_list)
    for row_data in value_list:
        worksheet.append(row_data)
    
    if query_type=="log_count":
        worksheet.append(['Total',total_amount])
    elif query_type=="error_count":
       worksheet.append(['Total',total_error_count])
    
    
 
    # Get the index of the last updated row
    last_row_index = worksheet.max_row

    # Make all cells in the last updated row bold
    for cell in worksheet[last_row_index]:
        cell.font = Font(bold=True)
       

        

    #save the chnages
    workbook.save(blob_name)

    # Upload the modified Excel file back to Azure Storage, replacing the original file
    with open(blob_name, "rb") as data:
        blob_client.upload_blob(data, overwrite=True)

    print("Data appended successfully.")
     

# execute_queries(log_count_queries,error_count_queries,"lock-error-count.xlsx")